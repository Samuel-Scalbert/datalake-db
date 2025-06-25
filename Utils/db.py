import os
import json
from tqdm import tqdm
from openpyxl import load_workbook

def check_or_create_collection(db, collection_name, collection_type='Collection'):
    """
    Checks if a collection exists in the database. If not, creates the collection.
    :param db: Database connection
    :param collection_name: Name of the collection
    :param collection_type: Type of the collection ('Collection' or 'Edges')
    :return: The collection object
    """
    if db.hasCollection(collection_name):
        return db[collection_name]
    else:
        db.createCollection(collection_type, name=collection_name)
        return db[collection_name]

def duplicates_JSON(lst):
    seen = set()
    duplicates = []

    for item in lst:
        item_hashable = str(item)
        if item_hashable in seen:
            duplicates.append(item)
        else:
            seen.add(item_hashable)

    return duplicates

def insert_json_db(data_path_json,db):
    list_errors = []
    db.dropAllCollections()

    workbook = load_workbook(filename='./app/data/Logiciels_Blacklist_et_autres_remarques.xlsx')
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    blacklist = []
    for row in data[1:]:
        blacklist.append(row[0])

    # Create or retrieve collections
    documents_collection = check_or_create_collection(db, 'documents')
    softwares_collection = check_or_create_collection(db, 'softwares')

    # Create or retrieve edge collections
    doc_soft_edge = check_or_create_collection(db, 'edge_doc_to_software', 'Edges')

    data_json_files = os.listdir(data_path_json)

    for file_path in tqdm(data_json_files[0:]):
        file_name = os.path.join(data_path_json, file_path)
        data_json_get_document = {}

        # FILE_ID -----------------------------------------------------

        data_json_get_document['file_hal_id'] = file_path.replace('.software.json','')
        document_document = documents_collection.createDocument(data_json_get_document)
        document_document.save()

        # SOFTWARE -----------------------------------------------------
        if file_path in data_json_files:
            with open(file_name, 'r') as json_file:
                data_json = json.load(json_file)
                data_json_get_mentions = data_json.get("mentions")

                # Remove duplicates
                for elm in duplicates_JSON(data_json_get_mentions):
                    data_json_get_mentions.remove(elm)

                # Process each mention
                for mention in data_json_get_mentions:
                    if mention['software-name']['normalizedForm'] not in blacklist:
                        mention['software_name'] = mention.pop('software-name')
                        mention['software_type'] = mention.pop('software-type')
                        software_document = softwares_collection.createDocument(mention)
                        software_document.save()

                        # Create edge from document to software
                        edge_doc_soft = doc_soft_edge.createEdge()
                        edge_doc_soft['_from'] = document_document._id
                        edge_doc_soft['_to'] = software_document._id
                        edge_doc_soft.save()

    if len(list_errors) > 0:
       print(list_errors)